from __future__ import annotations

import csv
import json
import os
import tempfile
from dataclasses import is_dataclass, asdict
from typing import Any, Iterable, Mapping
from core.scraper.request_params import normalize_params

PREVIEW_PREFERRED_COLUMNS: list[str] = [
    "task_id",
    "url",
    "final_url",
    "status_code",
    "title",
    "content_len",
    "request_ms",
    "redirects",
    "candidates_total",
    "candidates_xss",
    "candidates_sqli",
    "candidates_lfi",
    "candidates_ssrf",
    "candidates_types_present",
    "candidates_max_confidence",
    "findings_total",
    "findings_xss",
    "findings_sqli",
    "findings_lfi",
    "findings_ssrf",
    "findings_max_confidence",
    "findings_max_priority",
    "findings_with_baseline_hash",
    "findings_with_request_context",
    "findings_with_response_context",
    "findings_with_primary_evidence",
    "findings_types_present",
    "findings_priorities_present",
    "findings_confirmed_total",
    "findings_replay_ready_total",
    "findings_unique_replay_keys",
    "findings_unique_artifact_ids",
    "replay_groups_total",
    "replay_groups_replay_ready_total",
    "replay_groups_unique_targets",
    "replay_groups_max_group_size",
    "replay_groups_with_baseline_hash",
    "replay_groups_types_present",
    "replay_manifest_total",
    "replay_manifest_ready_total",
    "replay_manifest_with_headers",
    "replay_manifest_with_cookie_path",
    "replay_manifest_with_timeout",
    "replay_manifest_with_baseline_hash",
    "replay_manifest_unique_targets",
    "replay_manifest_types_present",
    "validation_plan_total",
    "validation_plan_ready_total",
    "validation_plan_with_param_targets",
    "validation_plan_reflection_checks",
    "validation_plan_error_pattern_checks",
    "validation_plan_status_diff_checks",
    "validation_plan_types_present",
    "validation_plan_checks_present",
    "validation_plan_total_checks",
    "validation_plan_unique_checks_present",
    "validation_plan_plans_with_checks",
    "validation_plan_plans_without_checks",
    "validation_plan_safe_mode_total",
    "validation_plan_candidate_targets_total",
    "validation_plan_unique_candidate_targets",
    "validation_plan_candidate_to_check_density",
    "validation_plan_avg_checks_per_plan",
    "validation_plan_avg_candidate_targets_per_plan",
    "validation_plan_ready_candidate_targets_total",
    "validation_plan_ready_candidate_targets_unique",
    "validation_plan_ready_candidate_coverage_ratio",
    "validation_plan_evidence_levels_present",
    "validation_plan_methods_present",
    "validation_plan_unique_targets",
    "validation_plan_target_sources_present",
    "validation_plan_targets_from_candidate_url",
    "validation_plan_targets_from_final_url",
    "validation_plan_targets_from_request_url",
    "validation_plan_targets_from_discovery_base_url",
    "validation_plan_targets_from_unknown",
    "validation_plan_total_check_plan_items",
    "validation_plan_ready_check_plan_items",
    "validation_plan_reflection_probe_items",
    "validation_plan_error_probe_items",
    "validation_plan_status_probe_items",
    "validation_plan_param_required_check_items",
    "validation_plan_non_param_check_items",
    "validation_plan_payload_families_present",
    "validation_plan_check_types_present",
    "validation_plan_plans_with_check_plan",
    "validation_plan_plans_without_check_plan",
    "validation_plan_avg_check_plan_items_per_plan",
    "validation_plan_plans_requiring_body_preview",
    "validation_plan_plans_requiring_status_code",
    "validation_plan_plans_requiring_body_hash",
    "validation_plan_plans_requiring_content_type",
    "validation_plan_check_plan_items_requiring_body_preview",
    "validation_plan_check_plan_items_requiring_status_code",
    "validation_plan_check_plan_items_requiring_body_hash",
    "validation_plan_check_plan_items_requiring_content_type",
    "validation_plan_comparison_modes_present",
    "validation_plan_reflection_compare_items",
    "validation_plan_error_pattern_compare_items",
    "validation_plan_status_compare_items",
    "validation_plan_baseline_field_body_preview_total",
    "validation_plan_baseline_field_status_code_total",
    "validation_plan_baseline_field_body_hash_total",
    "validation_plan_baseline_field_content_type_total",
    "validation_plan_execution_ready_check_plan_items",
    "validation_plan_blocked_check_plan_items",
    "validation_plan_plans_baseline_inputs_complete",
    "validation_plan_plans_with_missing_baseline_fields",
    "validation_plan_missing_body_preview_items",
    "validation_plan_missing_status_code_items",
    "validation_plan_missing_body_hash_items",
    "validation_plan_missing_content_type_items",
    "validation_plan_missing_baseline_fields_present",
    "validation_plan_plans_blocked_by_body_preview",
    "validation_plan_plans_blocked_by_status_code",
    "validation_plan_plans_blocked_by_body_hash",
    "validation_plan_plans_blocked_by_content_type",
    "validation_plan_execution_ready_ratio",
    "validation_plan_baseline_completeness_ratio",
    "validation_plan_parameterized_check_plan_items",
    "validation_plan_endpoint_level_check_plan_items",
    "validation_plan_plans_with_parameterized_checks",
    "validation_plan_plans_with_endpoint_level_checks",
    "validation_plan_total_effective_targets",
    "validation_plan_avg_effective_targets_per_plan",
    "validation_plan_execution_modes_present",
    "validation_plan_surface_types_present",
    "validation_plan_param_only_items",
    "validation_plan_endpoint_only_items",
    "validation_plan_hybrid_items",
    "validation_plan_unavailable_surface_items",
    "validation_plan_plans_param_only",
    "validation_plan_plans_endpoint_only",
    "validation_plan_plans_hybrid",
    "validation_plan_plans_with_unavailable_surface",
    "validation_plan_parameterized_ratio",
    "validation_plan_endpoint_level_ratio",
    "validation_plan_ready_param_items",
    "validation_plan_ready_endpoint_items",
    "validation_plan_blocked_param_items",
    "validation_plan_blocked_endpoint_items",
    "validation_plan_unavailable_items",
    "validation_plan_execution_lanes_present",
    "validation_plan_highest_priority_plans",
    "validation_plan_ready_queue_total",
    "validation_plan_blocked_queue_total",
    "validation_plan_avg_queue_size_per_plan",
    "validation_plan_avg_ready_queue_per_plan",
    "validation_plan_avg_blocked_queue_per_plan",
    "validation_plan_plans_primary_ready_param",
    "validation_plan_plans_primary_ready_endpoint",
    "validation_plan_plans_primary_blocked_param",
    "validation_plan_plans_primary_blocked_endpoint",
    "validation_plan_plans_primary_unavailable",
    "validation_plan_execution_priority_min",
    "validation_plan_execution_priority_max",
    "validation_plan_param_replace_items",
    "validation_plan_endpoint_compare_items",
    "validation_plan_unavailable_mutation_items",
    "validation_plan_mutating_check_plan_items",
    "validation_plan_baseline_only_check_plan_items",
    "validation_plan_mutation_ready_check_plan_items",
    "validation_plan_mutation_strategies_present",
    "validation_plan_plans_with_mutating_checks",
    "validation_plan_plans_with_baseline_only_checks",
    "validation_plan_total_mutation_slots",
    "validation_plan_avg_mutation_slots_per_plan",
    "validation_plan_mutation_ready_ratio",
    "validation_plan_plans_primary_param_replace",
    "validation_plan_plans_primary_endpoint_compare",
    "validation_plan_validator_jobs_total",
    "validation_plan_validator_jobs_ready",
    "validation_plan_validator_jobs_blocked",
    "validation_plan_validator_jobs_unavailable",
    "validation_plan_validator_job_types_present",
    "validation_plan_safe_validation_jobs",
    "validation_plan_blocked_validation_jobs",
    "validation_plan_unavailable_validation_jobs",
    "validation_plan_plans_with_ready_validator_jobs",
    "validation_plan_plans_with_blocked_validator_jobs",
    "validation_plan_plans_with_unavailable_validator_jobs",
    "validation_plan_validator_job_ready_ratio",
    "validation_plan_unique_validator_job_ids",
    "validation_plan_avg_validator_jobs_per_plan",
    "validation_plan_avg_ready_validator_jobs_per_plan",
    "validation_plan_plans_primary_safe_validation",
    "validation_plan_plans_primary_blocked_validation",
    "validation_plan_plans_primary_unavailable_validation",
    "validator_queue_total",
    "validator_queue_dispatch_ready_total",
    "validator_queue_jobs_total",
    "validator_queue_jobs_ready",
    "validator_queue_jobs_blocked",
    "validator_queue_jobs_unavailable",
    "validator_queue_unique_targets",
    "validator_queue_methods_present",
    "validator_queue_target_sources_present",
    "validator_queue_validator_job_types_present",
    "validator_queue_compare_modes_present",
    "validator_queue_mutation_strategies_present",
    "validator_queue_execution_lanes_present",
    "validator_queue_max_queue_size",
    "validator_queue_min_queue_size",
    "validator_queue_avg_queue_size",
    "validator_queue_ready_queue_ratio",
    "validator_queue_validator_job_ready_ratio",
    "validator_queue_primary_ready_queues",
    "validator_queue_unique_validator_job_ids",
    "validator_queue_ready_only_queues",
    "validator_queue_mixed_queues",
    "validator_queue_blocked_only_queues",
    "validator_queue_unavailable_only_queues",
    "validator_queue_empty_queues",
    "validator_queue_dispatch_jobs_total",
    "validator_queue_blocked_jobs_total",
    "validator_queue_unavailable_jobs_total",
    "validator_queue_fully_dispatchable_queues",
    "validator_queue_partially_dispatchable_queues",
    "validator_queue_primary_dispatchable_queues",
    "validator_queue_dispatch_modes_present",
    "validator_queue_avg_dispatch_jobs_per_queue",
    "validator_queue_avg_blocked_jobs_per_queue",
    "validator_queue_avg_unavailable_jobs_per_queue",
    "validator_queue_fully_dispatchable_ratio",
    "validator_queue_dispatch_job_ratio",
]

PREVIEW_HIDDEN_RAW_FIELDS: set[str] = {
    "candidates",
    "candidates_summary",
    "finding_artifacts",
    "discovery",
    "fingerprint",
    "forms",
    "headers",
    "timings",
    "parameter_intelligence",
    "request_recipe",
    "response_snapshot",
    "replay_groups",
    "replay_manifest",
    "validation_plan",
    "validator_queue",
}


# ---- Публичное API --------------------------------------------------------


def task_to_record(task_or_payload: Any) -> dict[str, Any]:
    """
    Приводит объект задачи (Task) ИЛИ "сырой" payload/dict к единому плоскому словарю (record).
    В record включаем самые полезные поля для анализа/экспорта.

    Поддерживаемые источники:
      - task объект: ожидаются атрибуты .id, .url, .method, .params, .result
      - dict payload: берём как есть (как будто это .result)
    """
    # 1) Вытащим task/result максимально безопасно
    if isinstance(task_or_payload, Mapping):
        task_id = task_or_payload.get("task_id") or task_or_payload.get("id") or ""
        url = task_or_payload.get("url") or ""
        method = _str_or(task_or_payload.get("method"), "")
        params = task_or_payload.get("params") or {}
        result = task_or_payload
    else:
        # объект задачи
        task = task_or_payload
        task_id = getattr(task, "id", "") or ""
        url = getattr(task, "url", "") or ""
        method = _str_or(getattr(task, "method", None), "")
        params = getattr(task, "params", {}) or {}
        result = getattr(task, "result", {}) or {}

        # dataclass params → dict
        if is_dataclass(params):
            params = asdict(params)

        # 2) Достаём поля из params/result
    normalized_params = normalize_params(params if isinstance(params, Mapping) else {})

    user_agent = _str_or(normalized_params.get("user_agent"), "")
    proxy = _str_or(normalized_params.get("proxy"), "")
    timeout = normalized_params.get("timeout")
    retries = normalized_params.get("retries")
    headers_req = normalized_params.get("headers", {})


    final_url = _str_or(_deep_get(result, "final_url"), "") or _str_or(result.get("url"), url)
    status_code = _deep_get(result, "status_code")
    content_len = _deep_get(result, "content_len")
    request_ms = _deep_get(result, "timings", "request_ms")
    title = _str_or(_deep_get(result, "title"), "")

    redirect_chain = _deep_get(result, "redirect_chain") or []
    redirects = None
    if isinstance(redirect_chain, (list, tuple)):
        redirects = len(redirect_chain)

    headers_resp = _deep_get(result, "headers")

    # 3) Сформируем export-friendly record без потери полезных derived полей из payload.
    #    Важно: начинаем с payload, чтобы preview/export использовали общий нормализованный источник.
    record: dict[str, Any] = dict(result) if isinstance(result, Mapping) else {}

    record.setdefault("task_id", task_id)
    record.setdefault("url", url)
    record["final_url"] = final_url
    record["method"] = method or _str_or(record.get("method"), "")
    record["status_code"] = status_code
    record["content_len"] = content_len
    record["request_ms"] = request_ms
    record["redirects"] = redirects
    record["title"] = title

    # Параметры запроса
    record["user_agent"] = user_agent
    record["proxy"] = proxy
    record["timeout"] = timeout
    record["retries"] = retries

    # Заголовки (req/resp) — как JSON-строки для CSV/XLSX
    record["headers_request"] = _maybe_json_string(headers_req)
    record["headers_response"] = _maybe_json_string(headers_resp)

    # Полный сырый result — удобно для JSON-экспорта без потерь:
    record["_raw_result"] = result

    record.update(derive_candidate_summary_fields(result))
    record.update(derive_finding_artifact_summary_fields(result))
    record.update(derive_replay_group_summary_fields(result))
    record.update(derive_replay_manifest_summary_fields(result))
    record.update(derive_validation_plan_summary_fields(result))
    record.update(derive_validator_queue_summary_fields(result))
    record.update(derive_request_recipe_summary_fields(result))
    record.update(derive_response_snapshot_summary_fields(result))

    return record


def derive_candidate_summary_fields(result: Any) -> dict[str, Any]:
    """
    Возвращает export-friendly candidate-derived summary fields
    из raw result без модификации исходного payload.
    """
    summary_defaults: dict[str, Any] = {
        "candidates_total": 0,
        "candidates_xss": 0,
        "candidates_sqli": 0,
        "candidates_lfi": 0,
        "candidates_ssrf": 0,
        "candidates_types_present": "",
        "candidates_max_confidence": "",
    }
    if not isinstance(result, Mapping):
        return summary_defaults

    out = dict(summary_defaults)
    out.update(_extract_candidate_counts(result))
    out.update(_extract_candidate_debug_fields(result))
    return out


def derive_response_snapshot_summary_fields(result: Any) -> dict[str, Any]:
    """
    Возвращает compact/flat поля из response_snapshot
    без модификации исходного payload.
    """
    defaults: dict[str, Any] = {
        "response_status_code": None,
        "response_content_type": "",
        "response_content_length": None,
        "response_body_hash": "",
        "response_has_body_preview": 0,
    }
    if not isinstance(result, Mapping):
        return defaults

    snapshot = result.get("response_snapshot")
    if not isinstance(snapshot, Mapping):
        return defaults

    body_preview = snapshot.get("body_preview")
    has_body_preview = 1 if isinstance(body_preview, str) and body_preview.strip() else 0
    status_code = snapshot.get("status_code")
    content_length = snapshot.get("content_length")

    return {
        "response_status_code": _to_int_count(status_code, None),
        "response_content_type": _str_or(snapshot.get("content_type"), ""),
        "response_content_length": _to_int_count(content_length, None),
        "response_body_hash": _str_or(snapshot.get("body_hash"), ""),
        "response_has_body_preview": has_body_preview,
    }


def derive_finding_artifact_summary_fields(result: Any) -> dict[str, Any]:
    defaults: dict[str, Any] = {
        "findings_total": 0,
        "findings_xss": 0,
        "findings_sqli": 0,
        "findings_lfi": 0,
        "findings_ssrf": 0,
        "findings_max_confidence": "",
        "findings_max_priority": "",
        "findings_with_baseline_hash": 0,
        "findings_with_request_context": 0,
        "findings_with_response_context": 0,
        "findings_with_primary_evidence": 0,
        "findings_types_present": "",
        "findings_priorities_present": "",
        "findings_confirmed_total": 0,
        "findings_replay_ready_total": 0,
        "findings_unique_replay_keys": 0,
        "findings_unique_artifact_ids": 0,
    }
    if not isinstance(result, Mapping):
        return defaults

    finding_artifacts = result.get("finding_artifacts")
    if not isinstance(finding_artifacts, Mapping):
        return defaults

    summary = finding_artifacts.get("summary")
    if not isinstance(summary, Mapping):
        return defaults

    by_type = summary.get("by_type") if isinstance(summary.get("by_type"), Mapping) else {}
    types_present = summary.get("types_present")
    priorities_present = summary.get("priorities_present")
    return {
        "findings_total": _to_int_count(summary.get("total"), 0),
        "findings_xss": _to_int_count(by_type.get("xss_candidate"), 0),
        "findings_sqli": _to_int_count(by_type.get("sqli_candidate"), 0),
        "findings_lfi": _to_int_count(by_type.get("lfi_candidate"), 0),
        "findings_ssrf": _to_int_count(by_type.get("ssrf_candidate"), 0),
        "findings_max_confidence": _str_or(summary.get("max_confidence"), ""),
        "findings_max_priority": _str_or(summary.get("max_priority"), ""),
        "findings_with_baseline_hash": _to_int_count(summary.get("with_baseline_hash"), 0),
        "findings_with_request_context": _to_int_count(summary.get("with_request_context"), 0),
        "findings_with_response_context": _to_int_count(summary.get("with_response_context"), 0),
        "findings_with_primary_evidence": _to_int_count(summary.get("with_primary_evidence"), 0),
        "findings_types_present": _join_string_list(types_present),
        "findings_priorities_present": _join_string_list(priorities_present),
        "findings_confirmed_total": _to_int_count(summary.get("confirmed_total"), 0),
        "findings_replay_ready_total": _to_int_count(summary.get("replay_ready_total"), 0),
        "findings_unique_replay_keys": _to_int_count(summary.get("unique_replay_keys"), 0),
        "findings_unique_artifact_ids": _to_int_count(summary.get("unique_artifact_ids"), 0),
    }


def derive_request_recipe_summary_fields(result: Any) -> dict[str, Any]:
    """
    Возвращает compact/flat поля из request_recipe
    без модификации исходного payload.
    """
    defaults: dict[str, Any] = {
        "request_url": "",
        "request_method": "",
        "request_redirects": 0,
        "request_timeout": "",
        "request_payload_source": "",
        "request_cookie_path_present": False,
    }
    if not isinstance(result, Mapping):
        return defaults

    recipe = result.get("request_recipe")
    if not isinstance(recipe, Mapping):
        return defaults

    timeout = recipe.get("timeout")

    return {
        "request_url": _str_or(recipe.get("url"), ""),
        "request_method": _str_or(recipe.get("method"), ""),
        "request_redirects": _to_int_count(recipe.get("redirects"), 0),
        "request_timeout": "" if timeout is None else _scalar_friendly(timeout),
        "request_payload_source": _str_or(recipe.get("payload_source"), ""),
        "request_cookie_path_present": bool(_str_or(recipe.get("cookie_path"), "").strip()),
    }


def derive_replay_group_summary_fields(result: Any) -> dict[str, Any]:
    defaults: dict[str, Any] = {
        "replay_groups_total": 0,
        "replay_groups_replay_ready_total": 0,
        "replay_groups_unique_targets": 0,
        "replay_groups_max_group_size": 0,
        "replay_groups_with_baseline_hash": 0,
        "replay_groups_types_present": "",
    }
    if not isinstance(result, Mapping):
        return defaults

    replay_groups = result.get("replay_groups")
    if not isinstance(replay_groups, Mapping):
        return defaults

    summary = replay_groups.get("summary")
    if not isinstance(summary, Mapping):
        return defaults

    return {
        "replay_groups_total": _to_int_count(summary.get("total"), 0),
        "replay_groups_replay_ready_total": _to_int_count(summary.get("replay_ready_total"), 0),
        "replay_groups_unique_targets": _to_int_count(summary.get("unique_targets"), 0),
        "replay_groups_max_group_size": _to_int_count(summary.get("max_group_size"), 0),
        "replay_groups_with_baseline_hash": _to_int_count(summary.get("with_baseline_hash"), 0),
        "replay_groups_types_present": _join_string_list(summary.get("types_present")),
    }


def derive_replay_manifest_summary_fields(result: Any) -> dict[str, Any]:
    defaults: dict[str, Any] = {
        "replay_manifest_total": 0,
        "replay_manifest_ready_total": 0,
        "replay_manifest_with_headers": 0,
        "replay_manifest_with_cookie_path": 0,
        "replay_manifest_with_timeout": 0,
        "replay_manifest_with_baseline_hash": 0,
        "replay_manifest_unique_targets": 0,
        "replay_manifest_types_present": "",
    }
    if not isinstance(result, Mapping):
        return defaults

    replay_manifest = result.get("replay_manifest")
    if not isinstance(replay_manifest, Mapping):
        return defaults

    summary = replay_manifest.get("summary")
    if not isinstance(summary, Mapping):
        return defaults

    return {
        "replay_manifest_total": _to_int_count(summary.get("total"), 0),
        "replay_manifest_ready_total": _to_int_count(summary.get("ready_total"), 0),
        "replay_manifest_with_headers": _to_int_count(summary.get("with_headers"), 0),
        "replay_manifest_with_cookie_path": _to_int_count(summary.get("with_cookie_path"), 0),
        "replay_manifest_with_timeout": _to_int_count(summary.get("with_timeout"), 0),
        "replay_manifest_with_baseline_hash": _to_int_count(summary.get("with_baseline_hash"), 0),
        "replay_manifest_unique_targets": _to_int_count(summary.get("unique_targets"), 0),
        "replay_manifest_types_present": _join_string_list(summary.get("types_present")),
    }


def derive_validation_plan_summary_fields(result: Any) -> dict[str, Any]:
    defaults: dict[str, Any] = {
        "validation_plan_total": 0,
        "validation_plan_ready_total": 0,
        "validation_plan_with_param_targets": 0,
        "validation_plan_reflection_checks": 0,
        "validation_plan_error_pattern_checks": 0,
        "validation_plan_status_diff_checks": 0,
        "validation_plan_types_present": "",
        "validation_plan_checks_present": "",
        "validation_plan_total_checks": 0,
        "validation_plan_unique_checks_present": "",
        "validation_plan_plans_with_checks": 0,
        "validation_plan_plans_without_checks": 0,
        "validation_plan_safe_mode_total": 0,
        "validation_plan_candidate_targets_total": 0,
        "validation_plan_unique_candidate_targets": 0,
        "validation_plan_candidate_to_check_density": 0.0,
        "validation_plan_avg_checks_per_plan": 0.0,
        "validation_plan_avg_candidate_targets_per_plan": 0.0,
        "validation_plan_ready_candidate_targets_total": 0,
        "validation_plan_ready_candidate_targets_unique": 0,
        "validation_plan_ready_candidate_coverage_ratio": 0.0,
        "validation_plan_evidence_levels_present": "",
        "validation_plan_methods_present": "",
        "validation_plan_unique_targets": 0,
        "validation_plan_target_sources_present": "",
        "validation_plan_targets_from_candidate_url": 0,
        "validation_plan_targets_from_final_url": 0,
        "validation_plan_targets_from_request_url": 0,
        "validation_plan_targets_from_discovery_base_url": 0,
        "validation_plan_targets_from_unknown": 0,
        "validation_plan_total_check_plan_items": 0,
        "validation_plan_ready_check_plan_items": 0,
        "validation_plan_reflection_probe_items": 0,
        "validation_plan_error_probe_items": 0,
        "validation_plan_status_probe_items": 0,
        "validation_plan_param_required_check_items": 0,
        "validation_plan_non_param_check_items": 0,
        "validation_plan_payload_families_present": "",
        "validation_plan_check_types_present": "",
        "validation_plan_plans_with_check_plan": 0,
        "validation_plan_plans_without_check_plan": 0,
        "validation_plan_avg_check_plan_items_per_plan": 0.0,
        "validation_plan_plans_requiring_body_preview": 0,
        "validation_plan_plans_requiring_status_code": 0,
        "validation_plan_plans_requiring_body_hash": 0,
        "validation_plan_plans_requiring_content_type": 0,
        "validation_plan_check_plan_items_requiring_body_preview": 0,
        "validation_plan_check_plan_items_requiring_status_code": 0,
        "validation_plan_check_plan_items_requiring_body_hash": 0,
        "validation_plan_check_plan_items_requiring_content_type": 0,
        "validation_plan_comparison_modes_present": "",
        "validation_plan_reflection_compare_items": 0,
        "validation_plan_error_pattern_compare_items": 0,
        "validation_plan_status_compare_items": 0,
        "validation_plan_baseline_field_body_preview_total": 0,
        "validation_plan_baseline_field_status_code_total": 0,
        "validation_plan_baseline_field_body_hash_total": 0,
        "validation_plan_baseline_field_content_type_total": 0,
        "validation_plan_execution_ready_check_plan_items": 0,
        "validation_plan_blocked_check_plan_items": 0,
        "validation_plan_plans_baseline_inputs_complete": 0,
        "validation_plan_plans_with_missing_baseline_fields": 0,
        "validation_plan_missing_body_preview_items": 0,
        "validation_plan_missing_status_code_items": 0,
        "validation_plan_missing_body_hash_items": 0,
        "validation_plan_missing_content_type_items": 0,
        "validation_plan_missing_baseline_fields_present": "",
        "validation_plan_plans_blocked_by_body_preview": 0,
        "validation_plan_plans_blocked_by_status_code": 0,
        "validation_plan_plans_blocked_by_body_hash": 0,
        "validation_plan_plans_blocked_by_content_type": 0,
        "validation_plan_execution_ready_ratio": 0.0,
        "validation_plan_baseline_completeness_ratio": 0.0,
        "validation_plan_parameterized_check_plan_items": 0,
        "validation_plan_endpoint_level_check_plan_items": 0,
        "validation_plan_plans_with_parameterized_checks": 0,
        "validation_plan_plans_with_endpoint_level_checks": 0,
        "validation_plan_total_effective_targets": 0,
        "validation_plan_avg_effective_targets_per_plan": 0.0,
        "validation_plan_execution_modes_present": "",
        "validation_plan_surface_types_present": "",
        "validation_plan_param_only_items": 0,
        "validation_plan_endpoint_only_items": 0,
        "validation_plan_hybrid_items": 0,
        "validation_plan_unavailable_surface_items": 0,
        "validation_plan_plans_param_only": 0,
        "validation_plan_plans_endpoint_only": 0,
        "validation_plan_plans_hybrid": 0,
        "validation_plan_plans_with_unavailable_surface": 0,
        "validation_plan_parameterized_ratio": 0.0,
        "validation_plan_endpoint_level_ratio": 0.0,
        "validation_plan_ready_param_items": 0,
        "validation_plan_ready_endpoint_items": 0,
        "validation_plan_blocked_param_items": 0,
        "validation_plan_blocked_endpoint_items": 0,
        "validation_plan_unavailable_items": 0,
        "validation_plan_execution_lanes_present": "",
        "validation_plan_highest_priority_plans": 0,
        "validation_plan_ready_queue_total": 0,
        "validation_plan_blocked_queue_total": 0,
        "validation_plan_avg_queue_size_per_plan": 0.0,
        "validation_plan_avg_ready_queue_per_plan": 0.0,
        "validation_plan_avg_blocked_queue_per_plan": 0.0,
        "validation_plan_plans_primary_ready_param": 0,
        "validation_plan_plans_primary_ready_endpoint": 0,
        "validation_plan_plans_primary_blocked_param": 0,
        "validation_plan_plans_primary_blocked_endpoint": 0,
        "validation_plan_plans_primary_unavailable": 0,
        "validation_plan_execution_priority_min": 0,
        "validation_plan_execution_priority_max": 0,
        "validation_plan_param_replace_items": 0,
        "validation_plan_endpoint_compare_items": 0,
        "validation_plan_unavailable_mutation_items": 0,
        "validation_plan_mutating_check_plan_items": 0,
        "validation_plan_baseline_only_check_plan_items": 0,
        "validation_plan_mutation_ready_check_plan_items": 0,
        "validation_plan_mutation_strategies_present": "",
        "validation_plan_plans_with_mutating_checks": 0,
        "validation_plan_plans_with_baseline_only_checks": 0,
        "validation_plan_total_mutation_slots": 0,
        "validation_plan_avg_mutation_slots_per_plan": 0.0,
        "validation_plan_mutation_ready_ratio": 0.0,
        "validation_plan_plans_primary_param_replace": 0,
        "validation_plan_plans_primary_endpoint_compare": 0,
        "validation_plan_validator_jobs_total": 0,
        "validation_plan_validator_jobs_ready": 0,
        "validation_plan_validator_jobs_blocked": 0,
        "validation_plan_validator_jobs_unavailable": 0,
        "validation_plan_validator_job_types_present": "",
        "validation_plan_safe_validation_jobs": 0,
        "validation_plan_blocked_validation_jobs": 0,
        "validation_plan_unavailable_validation_jobs": 0,
        "validation_plan_plans_with_ready_validator_jobs": 0,
        "validation_plan_plans_with_blocked_validator_jobs": 0,
        "validation_plan_plans_with_unavailable_validator_jobs": 0,
        "validation_plan_validator_job_ready_ratio": 0.0,
        "validation_plan_unique_validator_job_ids": 0,
        "validation_plan_avg_validator_jobs_per_plan": 0.0,
        "validation_plan_avg_ready_validator_jobs_per_plan": 0.0,
        "validation_plan_plans_primary_safe_validation": 0,
        "validation_plan_plans_primary_blocked_validation": 0,
        "validation_plan_plans_primary_unavailable_validation": 0,
    }
    if not isinstance(result, Mapping):
        return defaults

    validation_plan = result.get("validation_plan")
    if not isinstance(validation_plan, Mapping):
        return defaults

    summary = validation_plan.get("summary")
    if not isinstance(summary, Mapping):
        return defaults

    return {
        "validation_plan_total": _to_int_count(summary.get("total"), 0),
        "validation_plan_ready_total": _to_int_count(summary.get("ready_total"), 0),
        "validation_plan_with_param_targets": _to_int_count(summary.get("with_param_targets"), 0),
        "validation_plan_reflection_checks": _to_int_count(summary.get("reflection_checks"), 0),
        "validation_plan_error_pattern_checks": _to_int_count(summary.get("error_pattern_checks"), 0),
        "validation_plan_status_diff_checks": _to_int_count(summary.get("status_diff_checks"), 0),
        "validation_plan_types_present": _join_string_list(summary.get("types_present")),
        "validation_plan_checks_present": _join_string_list(summary.get("checks_present")),
        "validation_plan_total_checks": _to_int_count(summary.get("total_checks"), 0),
        "validation_plan_unique_checks_present": _join_string_list(summary.get("unique_checks_present")),
        "validation_plan_plans_with_checks": _to_int_count(summary.get("plans_with_checks"), 0),
        "validation_plan_plans_without_checks": _to_int_count(summary.get("plans_without_checks"), 0),
        "validation_plan_safe_mode_total": _to_int_count(summary.get("plans_safe_mode_total"), 0),
        "validation_plan_candidate_targets_total": _to_int_count(summary.get("candidate_targets_total"), 0),
        "validation_plan_unique_candidate_targets": _to_int_count(summary.get("unique_candidate_targets"), 0),
        "validation_plan_candidate_to_check_density": float(summary.get("candidate_to_check_density") or 0.0),
        "validation_plan_avg_checks_per_plan": float(summary.get("avg_checks_per_plan") or 0.0),
        "validation_plan_avg_candidate_targets_per_plan": float(
            summary.get("avg_candidate_targets_per_plan") or 0.0
        ),
        "validation_plan_ready_candidate_targets_total": _to_int_count(
            summary.get("ready_candidate_targets_total"), 0
        ),
        "validation_plan_ready_candidate_targets_unique": _to_int_count(
            summary.get("ready_candidate_targets_unique"), 0
        ),
        "validation_plan_ready_candidate_coverage_ratio": float(summary.get("ready_candidate_coverage_ratio") or 0.0),
        "validation_plan_evidence_levels_present": _join_string_list(summary.get("evidence_levels_present")),
        "validation_plan_methods_present": _join_string_list(summary.get("methods_present")),
        "validation_plan_unique_targets": _to_int_count(summary.get("unique_targets"), 0),
        "validation_plan_target_sources_present": _join_string_list(summary.get("target_sources_present")),
        "validation_plan_targets_from_candidate_url": _to_int_count(summary.get("targets_from_candidate_url"), 0),
        "validation_plan_targets_from_final_url": _to_int_count(summary.get("targets_from_final_url"), 0),
        "validation_plan_targets_from_request_url": _to_int_count(summary.get("targets_from_request_url"), 0),
        "validation_plan_targets_from_discovery_base_url": _to_int_count(summary.get("targets_from_discovery_base_url"), 0),
        "validation_plan_targets_from_unknown": _to_int_count(summary.get("targets_from_unknown"), 0),
        "validation_plan_total_check_plan_items": _to_int_count(summary.get("total_check_plan_items"), 0),
        "validation_plan_ready_check_plan_items": _to_int_count(summary.get("ready_check_plan_items"), 0),
        "validation_plan_reflection_probe_items": _to_int_count(summary.get("reflection_probe_items"), 0),
        "validation_plan_error_probe_items": _to_int_count(summary.get("error_probe_items"), 0),
        "validation_plan_status_probe_items": _to_int_count(summary.get("status_probe_items"), 0),
        "validation_plan_param_required_check_items": _to_int_count(summary.get("param_required_check_items"), 0),
        "validation_plan_non_param_check_items": _to_int_count(summary.get("non_param_check_items"), 0),
        "validation_plan_payload_families_present": _join_string_list(summary.get("payload_families_present")),
        "validation_plan_check_types_present": _join_string_list(summary.get("check_types_present")),
        "validation_plan_plans_with_check_plan": _to_int_count(summary.get("plans_with_check_plan"), 0),
        "validation_plan_plans_without_check_plan": _to_int_count(summary.get("plans_without_check_plan"), 0),
        "validation_plan_avg_check_plan_items_per_plan": float(summary.get("avg_check_plan_items_per_plan") or 0.0),
        "validation_plan_plans_requiring_body_preview": _to_int_count(summary.get("plans_requiring_body_preview"), 0),
        "validation_plan_plans_requiring_status_code": _to_int_count(summary.get("plans_requiring_status_code"), 0),
        "validation_plan_plans_requiring_body_hash": _to_int_count(summary.get("plans_requiring_body_hash"), 0),
        "validation_plan_plans_requiring_content_type": _to_int_count(summary.get("plans_requiring_content_type"), 0),
        "validation_plan_check_plan_items_requiring_body_preview": _to_int_count(
            summary.get("check_plan_items_requiring_body_preview"), 0
        ),
        "validation_plan_check_plan_items_requiring_status_code": _to_int_count(
            summary.get("check_plan_items_requiring_status_code"), 0
        ),
        "validation_plan_check_plan_items_requiring_body_hash": _to_int_count(
            summary.get("check_plan_items_requiring_body_hash"), 0
        ),
        "validation_plan_check_plan_items_requiring_content_type": _to_int_count(
            summary.get("check_plan_items_requiring_content_type"), 0
        ),
        "validation_plan_comparison_modes_present": _join_string_list(summary.get("comparison_modes_present")),
        "validation_plan_reflection_compare_items": _to_int_count(summary.get("reflection_compare_items"), 0),
        "validation_plan_error_pattern_compare_items": _to_int_count(summary.get("error_pattern_compare_items"), 0),
        "validation_plan_status_compare_items": _to_int_count(summary.get("status_compare_items"), 0),
        "validation_plan_baseline_field_body_preview_total": _to_int_count(
            summary.get("baseline_field_body_preview_total"), 0
        ),
        "validation_plan_baseline_field_status_code_total": _to_int_count(
            summary.get("baseline_field_status_code_total"), 0
        ),
        "validation_plan_baseline_field_body_hash_total": _to_int_count(
            summary.get("baseline_field_body_hash_total"), 0
        ),
        "validation_plan_baseline_field_content_type_total": _to_int_count(
            summary.get("baseline_field_content_type_total"), 0
        ),
        "validation_plan_execution_ready_check_plan_items": _to_int_count(
            summary.get("execution_ready_check_plan_items"), 0
        ),
        "validation_plan_blocked_check_plan_items": _to_int_count(summary.get("blocked_check_plan_items"), 0),
        "validation_plan_plans_baseline_inputs_complete": _to_int_count(
            summary.get("plans_baseline_inputs_complete"), 0
        ),
        "validation_plan_plans_with_missing_baseline_fields": _to_int_count(
            summary.get("plans_with_missing_baseline_fields"), 0
        ),
        "validation_plan_missing_body_preview_items": _to_int_count(summary.get("missing_body_preview_items"), 0),
        "validation_plan_missing_status_code_items": _to_int_count(summary.get("missing_status_code_items"), 0),
        "validation_plan_missing_body_hash_items": _to_int_count(summary.get("missing_body_hash_items"), 0),
        "validation_plan_missing_content_type_items": _to_int_count(summary.get("missing_content_type_items"), 0),
        "validation_plan_missing_baseline_fields_present": _join_string_list(
            summary.get("missing_baseline_fields_present")
        ),
        "validation_plan_plans_blocked_by_body_preview": _to_int_count(
            summary.get("plans_blocked_by_body_preview"), 0
        ),
        "validation_plan_plans_blocked_by_status_code": _to_int_count(
            summary.get("plans_blocked_by_status_code"), 0
        ),
        "validation_plan_plans_blocked_by_body_hash": _to_int_count(summary.get("plans_blocked_by_body_hash"), 0),
        "validation_plan_plans_blocked_by_content_type": _to_int_count(
            summary.get("plans_blocked_by_content_type"), 0
        ),
        "validation_plan_execution_ready_ratio": float(summary.get("execution_ready_ratio") or 0.0),
        "validation_plan_baseline_completeness_ratio": float(summary.get("baseline_completeness_ratio") or 0.0),
        "validation_plan_parameterized_check_plan_items": _to_int_count(
            summary.get("parameterized_check_plan_items"), 0
        ),
        "validation_plan_endpoint_level_check_plan_items": _to_int_count(
            summary.get("endpoint_level_check_plan_items"), 0
        ),
        "validation_plan_plans_with_parameterized_checks": _to_int_count(
            summary.get("plans_with_parameterized_checks"), 0
        ),
        "validation_plan_plans_with_endpoint_level_checks": _to_int_count(
            summary.get("plans_with_endpoint_level_checks"), 0
        ),
        "validation_plan_total_effective_targets": _to_int_count(summary.get("total_effective_targets"), 0),
        "validation_plan_avg_effective_targets_per_plan": float(
            summary.get("avg_effective_targets_per_plan") or 0.0
        ),
        "validation_plan_execution_modes_present": _join_string_list(summary.get("execution_modes_present")),
        "validation_plan_surface_types_present": _join_string_list(summary.get("surface_types_present")),
        "validation_plan_param_only_items": _to_int_count(summary.get("param_only_items"), 0),
        "validation_plan_endpoint_only_items": _to_int_count(summary.get("endpoint_only_items"), 0),
        "validation_plan_hybrid_items": _to_int_count(summary.get("hybrid_items"), 0),
        "validation_plan_unavailable_surface_items": _to_int_count(summary.get("unavailable_surface_items"), 0),
        "validation_plan_plans_param_only": _to_int_count(summary.get("plans_param_only"), 0),
        "validation_plan_plans_endpoint_only": _to_int_count(summary.get("plans_endpoint_only"), 0),
        "validation_plan_plans_hybrid": _to_int_count(summary.get("plans_hybrid"), 0),
        "validation_plan_plans_with_unavailable_surface": _to_int_count(
            summary.get("plans_with_unavailable_surface"), 0
        ),
        "validation_plan_parameterized_ratio": float(summary.get("parameterized_ratio") or 0.0),
        "validation_plan_endpoint_level_ratio": float(summary.get("endpoint_level_ratio") or 0.0),
        "validation_plan_ready_param_items": _to_int_count(summary.get("ready_param_items"), 0),
        "validation_plan_ready_endpoint_items": _to_int_count(summary.get("ready_endpoint_items"), 0),
        "validation_plan_blocked_param_items": _to_int_count(summary.get("blocked_param_items"), 0),
        "validation_plan_blocked_endpoint_items": _to_int_count(summary.get("blocked_endpoint_items"), 0),
        "validation_plan_unavailable_items": _to_int_count(summary.get("unavailable_items"), 0),
        "validation_plan_execution_lanes_present": _join_string_list(summary.get("execution_lanes_present")),
        "validation_plan_highest_priority_plans": _to_int_count(summary.get("highest_priority_plans"), 0),
        "validation_plan_ready_queue_total": _to_int_count(summary.get("ready_queue_total"), 0),
        "validation_plan_blocked_queue_total": _to_int_count(summary.get("blocked_queue_total"), 0),
        "validation_plan_avg_queue_size_per_plan": float(summary.get("avg_queue_size_per_plan") or 0.0),
        "validation_plan_avg_ready_queue_per_plan": float(summary.get("avg_ready_queue_per_plan") or 0.0),
        "validation_plan_avg_blocked_queue_per_plan": float(summary.get("avg_blocked_queue_per_plan") or 0.0),
        "validation_plan_plans_primary_ready_param": _to_int_count(summary.get("plans_primary_ready_param"), 0),
        "validation_plan_plans_primary_ready_endpoint": _to_int_count(
            summary.get("plans_primary_ready_endpoint"), 0
        ),
        "validation_plan_plans_primary_blocked_param": _to_int_count(
            summary.get("plans_primary_blocked_param"), 0
        ),
        "validation_plan_plans_primary_blocked_endpoint": _to_int_count(
            summary.get("plans_primary_blocked_endpoint"), 0
        ),
        "validation_plan_plans_primary_unavailable": _to_int_count(summary.get("plans_primary_unavailable"), 0),
        "validation_plan_execution_priority_min": _to_int_count(summary.get("execution_priority_min"), 0),
        "validation_plan_execution_priority_max": _to_int_count(summary.get("execution_priority_max"), 0),
        "validation_plan_param_replace_items": _to_int_count(summary.get("param_replace_items"), 0),
        "validation_plan_endpoint_compare_items": _to_int_count(summary.get("endpoint_compare_items"), 0),
        "validation_plan_unavailable_mutation_items": _to_int_count(summary.get("unavailable_mutation_items"), 0),
        "validation_plan_mutating_check_plan_items": _to_int_count(summary.get("mutating_check_plan_items"), 0),
        "validation_plan_baseline_only_check_plan_items": _to_int_count(
            summary.get("baseline_only_check_plan_items"), 0
        ),
        "validation_plan_mutation_ready_check_plan_items": _to_int_count(
            summary.get("mutation_ready_check_plan_items"), 0
        ),
        "validation_plan_mutation_strategies_present": _join_string_list(summary.get("mutation_strategies_present")),
        "validation_plan_plans_with_mutating_checks": _to_int_count(summary.get("plans_with_mutating_checks"), 0),
        "validation_plan_plans_with_baseline_only_checks": _to_int_count(
            summary.get("plans_with_baseline_only_checks"), 0
        ),
        "validation_plan_total_mutation_slots": _to_int_count(summary.get("total_mutation_slots"), 0),
        "validation_plan_avg_mutation_slots_per_plan": float(summary.get("avg_mutation_slots_per_plan") or 0.0),
        "validation_plan_mutation_ready_ratio": float(summary.get("mutation_ready_ratio") or 0.0),
        "validation_plan_plans_primary_param_replace": _to_int_count(summary.get("plans_primary_param_replace"), 0),
        "validation_plan_plans_primary_endpoint_compare": _to_int_count(
            summary.get("plans_primary_endpoint_compare"), 0
        ),
        "validation_plan_validator_jobs_total": _to_int_count(summary.get("validator_jobs_total"), 0),
        "validation_plan_validator_jobs_ready": _to_int_count(summary.get("validator_jobs_ready"), 0),
        "validation_plan_validator_jobs_blocked": _to_int_count(summary.get("validator_jobs_blocked"), 0),
        "validation_plan_validator_jobs_unavailable": _to_int_count(summary.get("validator_jobs_unavailable"), 0),
        "validation_plan_validator_job_types_present": _join_string_list(summary.get("validator_job_types_present")),
        "validation_plan_safe_validation_jobs": _to_int_count(summary.get("safe_validation_jobs"), 0),
        "validation_plan_blocked_validation_jobs": _to_int_count(summary.get("blocked_validation_jobs"), 0),
        "validation_plan_unavailable_validation_jobs": _to_int_count(summary.get("unavailable_validation_jobs"), 0),
        "validation_plan_plans_with_ready_validator_jobs": _to_int_count(
            summary.get("plans_with_ready_validator_jobs"), 0
        ),
        "validation_plan_plans_with_blocked_validator_jobs": _to_int_count(
            summary.get("plans_with_blocked_validator_jobs"), 0
        ),
        "validation_plan_plans_with_unavailable_validator_jobs": _to_int_count(
            summary.get("plans_with_unavailable_validator_jobs"), 0
        ),
        "validation_plan_validator_job_ready_ratio": float(summary.get("validator_job_ready_ratio") or 0.0),
        "validation_plan_unique_validator_job_ids": _to_int_count(summary.get("unique_validator_job_ids"), 0),
        "validation_plan_avg_validator_jobs_per_plan": float(summary.get("avg_validator_jobs_per_plan") or 0.0),
        "validation_plan_avg_ready_validator_jobs_per_plan": float(
            summary.get("avg_ready_validator_jobs_per_plan") or 0.0
        ),
        "validation_plan_plans_primary_safe_validation": _to_int_count(
            summary.get("plans_primary_safe_validation"), 0
        ),
        "validation_plan_plans_primary_blocked_validation": _to_int_count(
            summary.get("plans_primary_blocked_validation"), 0
        ),
        "validation_plan_plans_primary_unavailable_validation": _to_int_count(
            summary.get("plans_primary_unavailable_validation"), 0
        ),
    }


def derive_validator_queue_summary_fields(result: Any) -> dict[str, Any]:
    defaults: dict[str, Any] = {
        "validator_queue_total": 0,
        "validator_queue_dispatch_ready_total": 0,
        "validator_queue_jobs_total": 0,
        "validator_queue_jobs_ready": 0,
        "validator_queue_jobs_blocked": 0,
        "validator_queue_jobs_unavailable": 0,
        "validator_queue_unique_targets": 0,
        "validator_queue_methods_present": "",
        "validator_queue_target_sources_present": "",
        "validator_queue_validator_job_types_present": "",
        "validator_queue_compare_modes_present": "",
        "validator_queue_mutation_strategies_present": "",
        "validator_queue_execution_lanes_present": "",
        "validator_queue_max_queue_size": 0,
        "validator_queue_min_queue_size": 0,
        "validator_queue_avg_queue_size": 0.0,
        "validator_queue_ready_queue_ratio": 0.0,
        "validator_queue_validator_job_ready_ratio": 0.0,
        "validator_queue_primary_ready_queues": 0,
        "validator_queue_unique_validator_job_ids": 0,
        "validator_queue_ready_only_queues": 0,
        "validator_queue_mixed_queues": 0,
        "validator_queue_blocked_only_queues": 0,
        "validator_queue_unavailable_only_queues": 0,
        "validator_queue_empty_queues": 0,
        "validator_queue_dispatch_jobs_total": 0,
        "validator_queue_blocked_jobs_total": 0,
        "validator_queue_unavailable_jobs_total": 0,
        "validator_queue_fully_dispatchable_queues": 0,
        "validator_queue_partially_dispatchable_queues": 0,
        "validator_queue_primary_dispatchable_queues": 0,
        "validator_queue_dispatch_modes_present": "",
        "validator_queue_avg_dispatch_jobs_per_queue": 0.0,
        "validator_queue_avg_blocked_jobs_per_queue": 0.0,
        "validator_queue_avg_unavailable_jobs_per_queue": 0.0,
        "validator_queue_fully_dispatchable_ratio": 0.0,
        "validator_queue_dispatch_job_ratio": 0.0,
    }
    if not isinstance(result, Mapping):
        return defaults
    validator_queue = result.get("validator_queue")
    if not isinstance(validator_queue, Mapping):
        return defaults
    summary = validator_queue.get("summary")
    if not isinstance(summary, Mapping):
        return defaults
    return {
        "validator_queue_total": _to_int_count(summary.get("total"), 0),
        "validator_queue_dispatch_ready_total": _to_int_count(summary.get("dispatch_ready_total"), 0),
        "validator_queue_jobs_total": _to_int_count(summary.get("jobs_total"), 0),
        "validator_queue_jobs_ready": _to_int_count(summary.get("jobs_ready"), 0),
        "validator_queue_jobs_blocked": _to_int_count(summary.get("jobs_blocked"), 0),
        "validator_queue_jobs_unavailable": _to_int_count(summary.get("jobs_unavailable"), 0),
        "validator_queue_unique_targets": _to_int_count(summary.get("unique_targets"), 0),
        "validator_queue_methods_present": _join_string_list(summary.get("methods_present")),
        "validator_queue_target_sources_present": _join_string_list(summary.get("target_sources_present")),
        "validator_queue_validator_job_types_present": _join_string_list(summary.get("validator_job_types_present")),
        "validator_queue_compare_modes_present": _join_string_list(summary.get("compare_modes_present")),
        "validator_queue_mutation_strategies_present": _join_string_list(summary.get("mutation_strategies_present")),
        "validator_queue_execution_lanes_present": _join_string_list(summary.get("execution_lanes_present")),
        "validator_queue_max_queue_size": _to_int_count(summary.get("max_queue_size"), 0),
        "validator_queue_min_queue_size": _to_int_count(summary.get("min_queue_size"), 0),
        "validator_queue_avg_queue_size": float(summary.get("avg_queue_size") or 0.0),
        "validator_queue_ready_queue_ratio": float(summary.get("ready_queue_ratio") or 0.0),
        "validator_queue_validator_job_ready_ratio": float(summary.get("validator_job_ready_ratio") or 0.0),
        "validator_queue_primary_ready_queues": _to_int_count(summary.get("primary_ready_queues"), 0),
        "validator_queue_unique_validator_job_ids": _to_int_count(summary.get("unique_validator_job_ids"), 0),
        "validator_queue_ready_only_queues": _to_int_count(summary.get("ready_only_queues"), 0),
        "validator_queue_mixed_queues": _to_int_count(summary.get("mixed_queues"), 0),
        "validator_queue_blocked_only_queues": _to_int_count(summary.get("blocked_only_queues"), 0),
        "validator_queue_unavailable_only_queues": _to_int_count(summary.get("unavailable_only_queues"), 0),
        "validator_queue_empty_queues": _to_int_count(summary.get("empty_queues"), 0),
        "validator_queue_dispatch_jobs_total": _to_int_count(summary.get("dispatch_jobs_total"), 0),
        "validator_queue_blocked_jobs_total": _to_int_count(summary.get("blocked_jobs_total"), 0),
        "validator_queue_unavailable_jobs_total": _to_int_count(summary.get("unavailable_jobs_total"), 0),
        "validator_queue_fully_dispatchable_queues": _to_int_count(summary.get("fully_dispatchable_queues"), 0),
        "validator_queue_partially_dispatchable_queues": _to_int_count(
            summary.get("partially_dispatchable_queues"), 0
        ),
        "validator_queue_primary_dispatchable_queues": _to_int_count(summary.get("primary_dispatchable_queues"), 0),
        "validator_queue_dispatch_modes_present": _join_string_list(summary.get("dispatch_modes_present")),
        "validator_queue_avg_dispatch_jobs_per_queue": float(summary.get("avg_dispatch_jobs_per_queue") or 0.0),
        "validator_queue_avg_blocked_jobs_per_queue": float(summary.get("avg_blocked_jobs_per_queue") or 0.0),
        "validator_queue_avg_unavailable_jobs_per_queue": float(
            summary.get("avg_unavailable_jobs_per_queue") or 0.0
        ),
        "validator_queue_fully_dispatchable_ratio": float(summary.get("fully_dispatchable_ratio") or 0.0),
        "validator_queue_dispatch_job_ratio": float(summary.get("dispatch_job_ratio") or 0.0),
    }


def export(records: Iterable[Mapping[str, Any]], path: str, fmt: str = "csv") -> str:
    """
    Единая точка экспорта. Поддержка форматов: csv/json/xlsx
    - records: итератор словарей (результат task_to_record или аналогичный)
    - path: конечный путь (расширение можно не учитывать — приоритет у fmt)
    - fmt: 'csv' | 'json' | 'xlsx'

    Возвращает финальный путь к записанному файлу.
    """
    fmt = (fmt or "").lower().strip()
    if fmt not in {"csv", "json", "xlsx"}:
        raise ValueError(f"Unsupported export format: {fmt}")

    source_records = [dict(r) for r in records]
    # Нормализуем список к export-friendly плоскому виду (включая candidate summary поля).
    items = [task_to_record(r) for r in source_records]

    # Гарантируем существование директории
    _ensure_parent_dir(path)

    if fmt == "json":
        _to_json(items, path)
    else:
        tabular_items = normalize_preview_rows(source_records)
        columns = preview_column_order(tabular_items)
        if fmt == "csv":
            _to_csv(tabular_items, path, fieldnames=columns)
        else:
            _to_xlsx(tabular_items, path, fieldnames=columns)

    return path


# ---- Внутренние реализации -------------------------------------------------


def _to_json(items: list[dict[str, Any]], path: str) -> None:
    """
    JSON экспорт: сохраняем массив records, но поле '_raw_result' оставляем как есть (dict),
    чтобы не терять данные. Остальные поля уже плоские.
    """
    # Безопасная атомарная запись
    data = items
    _atomic_write_json(path, data, ensure_ascii=False, indent=2)


def _to_csv(items: list[dict[str, Any]], path: str, fieldnames: list[str] | None = None) -> None:
    """
    CSV экспорт: объединяем все ключи как заголовки,
    dict/list значения сериализуем в JSON-строку,
    кодировка utf-8-sig (Excel-friendly).
    """
    if not items:
        # создадим пустой CSV с нулевой строкой заголовков
        _atomic_write_text(path, "", encoding="utf-8-sig")
        return

    # Объединение ключей всех строк
    fieldnames = fieldnames or _union_keys(items)

    def row_gen():
        for it in items:
            yield {k: _scalarize(it.get(k)) for k in fieldnames}

    with _atomic_writer(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, dialect="excel", quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for row in row_gen():
            writer.writerow(row)


def _to_xlsx(items: list[dict[str, Any]], path: str, fieldnames: list[str] | None = None) -> None:
    """
    XLSX экспорт через openpyxl.
    dict/list → JSON-строки (как в CSV).
    """
    try:
        import openpyxl  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "Для экспорта в .xlsx требуется пакет 'openpyxl'. Установите: pip install openpyxl"
        ) from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "export"

    fieldnames = fieldnames or _union_keys(items)

    # Заголовки
    ws.append(list(fieldnames))

    # Строки
    for it in items:
        ws.append([_scalarize(it.get(k)) for k in fieldnames])

    # Простая авто-ширина по максимуму длины значения (ограничим до разумных пределов)
    MAX = 60
    for idx, key in enumerate(fieldnames, start=1):
        col = get_column_letter(idx)
        max_len = len(str(key))
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            val = row[0].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col].width = min(MAX, max_len + 2)

    # Атомарная запись файла
    _atomic_write_xlsx(wb, path)


# ---- Утилиты ---------------------------------------------------------------


def _str_or(value: Any, default: str) -> str:
    return value if isinstance(value, str) else default


def _to_int_count(value: Any, default: int = 0) -> int:
    try:
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, (int, float)):
            return int(value)
        if isinstance(value, str):
            return int(value.strip())
    except Exception:
        pass
    return default


def _join_string_list(value: Any) -> str:
    if not isinstance(value, list):
        return ""
    out: list[str] = []
    seen: set[str] = set()
    for item in value:
        if isinstance(item, str):
            clean = item.strip()
            if clean and clean not in seen:
                out.append(clean)
                seen.add(clean)
    return ",".join(out)


def _extract_candidate_counts(rec: Mapping[str, Any]) -> dict[str, int]:
    summary = rec.get("candidates_summary")
    if not isinstance(summary, Mapping):
        candidates = rec.get("candidates")
        if isinstance(candidates, Mapping):
            summary = candidates.get("summary")
        else:
            summary = None

    if not isinstance(summary, Mapping):
        return {}

    by_type = summary.get("by_type") if isinstance(summary.get("by_type"), Mapping) else {}
    return {
        "candidates_total": _to_int_count(summary.get("total"), 0),
        "candidates_xss": _to_int_count(by_type.get("xss_candidate"), 0),
        "candidates_sqli": _to_int_count(by_type.get("sqli_candidate"), 0),
        "candidates_lfi": _to_int_count(by_type.get("lfi_candidate"), 0),
        "candidates_ssrf": _to_int_count(by_type.get("ssrf_candidate"), 0),
    }


def _extract_candidate_debug_fields(rec: Mapping[str, Any]) -> dict[str, str]:
    candidates = rec.get("candidates")
    candidates_all = candidates.get("all") if isinstance(candidates, Mapping) else None
    if not isinstance(candidates_all, list):
        return {
            "candidates_types_present": "",
            "candidates_max_confidence": "",
        }

    type_map = {
        "xss_candidate": "xss",
        "sqli_candidate": "sqli",
        "lfi_candidate": "lfi",
        "ssrf_candidate": "ssrf",
    }
    type_order = ("xss_candidate", "sqli_candidate", "lfi_candidate", "ssrf_candidate")
    confidence_rank = {
        "low": 1,
        "medium": 2,
        "high": 3,
    }

    found_types = set()
    max_confidence_value = ""
    max_confidence_rank = 0

    for item in candidates_all:
        if not isinstance(item, Mapping):
            continue

        raw_type = item.get("type")
        if isinstance(raw_type, str) and raw_type in type_map:
            found_types.add(raw_type)

        raw_confidence = item.get("confidence")
        if isinstance(raw_confidence, str):
            conf = raw_confidence.strip().lower()
            rank = confidence_rank.get(conf, 0)
            if rank > max_confidence_rank:
                max_confidence_rank = rank
                max_confidence_value = conf

    return {
        "candidates_types_present": ",".join(type_map[t] for t in type_order if t in found_types),
        "candidates_max_confidence": max_confidence_value,
    }


def _deep_get(obj: Any, *keys: str) -> Any:
    cur = obj
    for k in keys:
        if isinstance(cur, Mapping) and k in cur:
            cur = cur[k]
        else:
            return None
    return cur


def _maybe_json_string(val: Any) -> Any:
    """Возвращает JSON-строку для dict/list, иначе оригинальное значение."""
    if isinstance(val, (dict, list, tuple)):
        try:
            return json.dumps(val, ensure_ascii=False)
        except Exception:
            return str(val)
    return val


def _scalarize(val: Any) -> Any:
    """
    Приводит значение к безопасному для CSV/XLSX виду:
    - dict/list → JSON-строка
    - None → ""
    - остальное → как есть
    """
    if val is None:
        return ""
    if isinstance(val, (dict, list, tuple)):
        try:
            return json.dumps(val, ensure_ascii=False)
        except Exception:
            return str(val)
    return val


def _union_keys(items: list[dict[str, Any]]) -> list[str]:
    keys: dict[str, None] = {}
    for it in items:
        for k in it.keys():
            keys.setdefault(k, None)
    return list(keys.keys())


def preview_column_order(items: list[Mapping[str, Any]]) -> list[str]:
    """
    Общий порядок колонок для Data Preview и CSV/XLSX export:
    preferred -> остальные (sorted).
    """
    keys = set()
    for it in items:
        keys.update(it.keys())
    rest = sorted(k for k in keys if k not in PREVIEW_PREFERRED_COLUMNS)
    return [k for k in PREVIEW_PREFERRED_COLUMNS if k in keys] + rest


def _tabular_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Фильтрует внутренние/сырые поля для CSV/XLSX экспорта.
    """
    hidden = {
        "_raw_result",
        "candidates",
        "candidates_summary",
        "replay_groups",
        "replay_manifest",
        "validation_plan",
        "validator_queue",
    }
    out: list[dict[str, Any]] = []
    for it in items:
        out.append({k: v for k, v in it.items() if k not in hidden})
    return out


def normalize_preview_rows(records: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    """
    Приводит записи к той же normalized schema, что используется в Data Preview table:
    - нормализует ключи;
    - добавляет derived candidate summary;
    - скрывает bulky/raw поля, которых нет в превью-таблице;
    - применяет compact-значения для колонок превью.
    """
    out: list[dict[str, Any]] = []
    for rec in records:
        if not isinstance(rec, Mapping):
            rec = {"value": rec}

        row: dict[str, Any] = {}
        for key, value in rec.items():
            nk = _normalize_preview_key(key)
            if nk in row:
                i = 2
                while f"{nk}#{i}" in row:
                    i += 1
                nk = f"{nk}#{i}"
            row[nk] = value

        row.update(derive_candidate_summary_fields(rec))
        row.update(derive_finding_artifact_summary_fields(rec))
        row.update(derive_replay_group_summary_fields(rec))
        row.update(derive_replay_manifest_summary_fields(rec))
        row.update(derive_validation_plan_summary_fields(rec))
        row.update(derive_validator_queue_summary_fields(rec))
        row.update(derive_request_recipe_summary_fields(rec))
        row.update(derive_response_snapshot_summary_fields(rec))

        for hidden_key in PREVIEW_HIDDEN_RAW_FIELDS:
            row.pop(hidden_key, None)

        compacted: dict[str, Any] = {}
        for key, value in row.items():
            compacted[key] = _compact_preview_value(key, value)

        out.append(compacted)
    return out


def _normalize_preview_key(key: Any) -> str:
    if key is None:
        return "__none__"
    norm = str(key).replace("\ufeff", "").strip()
    return norm if norm else "__empty__"


def _compact_preview_value(key: str, val: Any) -> Any:
    if key == "forms_summary" and isinstance(val, Mapping):
        fs = val
        return (
            f"total={fs.get('forms_total', 0)}, "
            f"unique={fs.get('forms_unique', fs.get('forms_total', 0))}, "
            f"inputs={fs.get('inputs_total', 0)}, "
            f"unique_inputs={fs.get('inputs_unique_total', fs.get('inputs_total', 0))}, "
            f"names={fs.get('unique_input_names', 0)}"
        )
    if key == "forms" and isinstance(val, list):
        return _compact_forms(val)
    return val


def _scalar_friendly(val: Any) -> Any:
    if val is None:
        return ""
    if isinstance(val, (str, int, float, bool)):
        return val
    if isinstance(val, (dict, list, tuple)):
        return _maybe_json_string(val)
    return str(val)


def _compact_forms(forms: list[Any]) -> str:
    if not forms:
        return "[]"
    parts: list[str] = []
    max_forms = 2
    for idx, form in enumerate(forms[:max_forms], 1):
        if not isinstance(form, Mapping):
            continue
        method = str(form.get("method") or "").upper()
        action = str(form.get("action") or "")
        action_short = action if len(action) <= 120 else action[:119] + "…"
        enctype = form.get("enctype") or ""
        inputs_count = form.get("inputs_count") or len(form.get("inputs", []) or [])
        has_file = 1 if form.get("has_file") else 0
        names = form.get("input_names") or []
        names_short = ", ".join((str(n) if n is not None else "") for n in names[:10])
        if len(names) > 10:
            names_short += ", …"
        parts.append(
            f"[{idx}] {method} {action_short} enctype={enctype} inputs={inputs_count} file={has_file} names={names_short}"
        )
    if len(forms) > max_forms:
        parts.append(f"... +{len(forms) - max_forms} more")
    out = " | ".join(parts)
    return out if len(out) <= 2000 else out[:1999] + "…"


def _ensure_parent_dir(path: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)


# ---- Атомарные записи ------------------------------------------------------


def _atomic_writer(path: str, mode: str, **kwargs):
    """
    Контекстный менеджер для атомарной записи текстовых файлов.
    Записываем во временный файл и затем os.replace → path.
    """
    tmp_dir = os.path.dirname(os.path.abspath(path)) or "."
    fd, tmp_path = tempfile.mkstemp(prefix=".tmp_", dir=tmp_dir)
    os.close(fd)

    class _CM:
        def __enter__(self_inner):
            self_inner._f = open(tmp_path, mode, **kwargs)
            return self_inner._f

        def __exit__(self_inner, exc_type, exc, tb):
            self_inner._f.close()
            if exc_type is None:
                os.replace(tmp_path, path)
            else:
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass

    return _CM()


def _atomic_write_text(path: str, text: str, encoding: str = "utf-8") -> None:
    with _atomic_writer(path, "w", encoding=encoding, newline="") as f:
        f.write(text)


def _atomic_write_json(path: str, data: Any, ensure_ascii: bool = False, indent: int | None = 2) -> None:
    with _atomic_writer(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=ensure_ascii, indent=indent)


def _atomic_write_xlsx(workbook: Any, path: str) -> None:
    # openpyxl сам пишет атомарно плохо, поэтому вручную во временный и replace
    tmp_dir = os.path.dirname(os.path.abspath(path)) or "."
    fd, tmp_path = tempfile.mkstemp(prefix=".tmp_", suffix=".xlsx", dir=tmp_dir)
    os.close(fd)
    try:
        workbook.save(tmp_path)
        os.replace(tmp_path, path)
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass
