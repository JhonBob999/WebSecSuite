# core/scraper/exporter.py
from __future__ import annotations

# === SECTION === Imports & Typing
import csv
import json
import os
from typing import Iterable, Any, Dict, List, Optional
from pathlib import Path
from datetime import datetime

# === SECTION === Constants
_HEADERS_OF_INTEREST = (
    "server",
    "content-type",
    "content-length",
    "date",
    "via",
    "x-powered-by",
)

# === SECTION === Helpers: Filesystem
def _ensure_dir(path: str) -> None:
    folder = os.path.dirname(path)
    if folder and not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)

# === SECTION === Helpers: Result normalization
def _stringify_redirect_chain(chain: Any) -> str:
    """
    Convert redirect chain into a compact human-readable string:
    301→https://a | 302→https://b
    Supports list[dict] or list[tuple] or list[str].
    """
    if not chain:
        return ""
    out: List[str] = []
    for hop in chain:
        code = ""
        url = ""
        if isinstance(hop, dict):
            code = str(hop.get("status") or hop.get("code") or hop.get("status_code") or "")
            url = str(hop.get("url") or hop.get("location") or "")
        elif isinstance(hop, (list, tuple)) and len(hop) >= 2:
            code, url = str(hop[0]), str(hop[1])
        else:
            url = str(hop)
        if url:
            out.append(f"{code}→{url}" if code else url)
    return " | ".join(out)

def _lower_keys(d: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not isinstance(d, dict):
        return {}
    return {str(k).lower(): v for k, v in d.items()}

def _flatten_result_for_table(result: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Produce a flat row with key metrics for CSV/XLSX.
    Keeps most-used fields and picks important headers.
    """
    row: Dict[str, Any] = {
        "result_status_code": "",
        "result_final_url": "",
        "result_title": "",
        "result_content_len": "",
        "result_request_ms": "",
        "result_total_ms": "",
        "result_redirects": "",
    }
    if not isinstance(result, dict):
        return row

    # basics
    row["result_status_code"] = result.get("status_code", "")
    row["result_final_url"] = result.get("final_url", "")
    row["result_title"] = result.get("title", "")
    row["result_content_len"] = result.get("content_len", "")

    # timings
    timings = result.get("timings", {}) or {}
    row["result_request_ms"] = timings.get("request_ms", "")
    row["result_total_ms"] = timings.get("total_ms", "")

    # redirects
    row["result_redirects"] = _stringify_redirect_chain(result.get("redirect_chain"))

    # headers of interest
    headers = _lower_keys(result.get("headers"))
    for hk in _HEADERS_OF_INTEREST:
        row[f"header_{hk}"] = headers.get(hk, "")

    return row

# === SECTION === Helpers: Task → Row
def _task_to_row(task: Any) -> Dict[str, Any]:
    """
    Convert task (object or dict-like) to a flat row for tabular exports.
    Expected task attributes: id, url, method, status, progress, retries, timeout, user_agent, proxy, result.
    """
    # base fields
    base: Dict[str, Any] = {
        "id": getattr(task, "id", "") if not isinstance(task, dict) else task.get("id", ""),
        "url": getattr(task, "url", "") if not isinstance(task, dict) else task.get("url", ""),
        "method": getattr(task, "method", "GET") if not isinstance(task, dict) else (task.get("method") or "GET"),
        "status": "",
        "progress": getattr(task, "progress", 0) if not isinstance(task, dict) else task.get("progress", 0),
        "retries": getattr(task, "retries", 0) if not isinstance(task, dict) else task.get("retries", 0),
        "timeout": getattr(task, "timeout", "") if not isinstance(task, dict) else task.get("timeout", ""),
        "user_agent": getattr(task, "user_agent", "") if not isinstance(task, dict) else task.get("user_agent", ""),
        "proxy": getattr(task, "proxy", "") if not isinstance(task, dict) else task.get("proxy", ""),
    }

    # status may be enum-like
    if isinstance(task, dict):
        status = task.get("status", "")
    else:
        status = getattr(task, "status", "")
    base["status"] = getattr(status, "value", str(status))

    # attach flattened result part
    result = getattr(task, "result", None) if not isinstance(task, dict) else task.get("result")
    base.update(_flatten_result_for_table(result))
    return base

# === SECTION === Exporters (CSV / JSON / XLSX)
def _export_csv(tasks: Iterable[Any], path: str) -> None:
    rows = [_task_to_row(t) for t in tasks]
    if not rows:
        raise ValueError("No data to export.")
    _ensure_dir(path)

    # RFC4180-friendly for Excel: add BOM so Excel picks UTF-8 automatically
    fieldnames = list(rows[0].keys())
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

def _export_json(tasks: Iterable[Any], path: str) -> None:
    """
    Export FULL JSON: base task fields + full `result` object (no flattening),
    чтобы не терять headers/redirect_chain и пр.
    """
    out: List[Dict[str, Any]] = []
    for t in tasks:
        item = {
            "id": getattr(t, "id", "") if not isinstance(t, dict) else t.get("id", ""),
            "url": getattr(t, "url", "") if not isinstance(t, dict) else t.get("url", ""),
            "method": getattr(t, "method", "GET") if not isinstance(t, dict) else (t.get("method") or "GET"),
            "status": "",
            "progress": getattr(t, "progress", 0) if not isinstance(t, dict) else t.get("progress", 0),
            "retries": getattr(t, "retries", 0) if not isinstance(t, dict) else t.get("retries", 0),
            "timeout": getattr(t, "timeout", "") if not isinstance(t, dict) else t.get("timeout", ""),
            "user_agent": getattr(t, "user_agent", "") if not isinstance(t, dict) else t.get("user_agent", ""),
            "proxy": getattr(t, "proxy", "") if not isinstance(t, dict) else t.get("proxy", ""),
        }
        status = getattr(t, "status", "") if not isinstance(t, dict) else t.get("status", "")
        item["status"] = getattr(status, "value", str(status))
        # full result
        item["result"] = getattr(t, "result", None) if not isinstance(t, dict) else t.get("result")
        out.append(item)

    _ensure_dir(path)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

def _export_xlsx(tasks: Iterable[Any], path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.worksheet import Worksheet
    except Exception as e:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl") from e

    rows = [_task_to_row(t) for t in tasks]
    if not rows:
        raise ValueError("No data to export.")
    _ensure_dir(path)

    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Scraper"

    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # Freeze header row & autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}1048576"

    # Simple auto-width heuristic
    for col_idx, h in enumerate(headers, start=1):
        max_len = max(len(h), *(len(str(row.get(h, ""))) for row in rows))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 80)

    wb.save(path)

# === SECTION === Public API
def export_tasks(tasks: Iterable[Any], fmt: str, path: str) -> str:
    """
    Export tasks into chosen format.
    fmt: "csv" | "json" | "xlsx"
    Returns the path.
    """
    fmt = (fmt or "").lower()
    if fmt not in {"csv", "json", "xlsx"}:
        raise ValueError(f"Unsupported format: {fmt}")

    if fmt == "csv":
        _export_csv(tasks, path)
    elif fmt == "json":
        _export_json(tasks, path)
    else:  # xlsx
        _export_xlsx(tasks, path)
    return path

def default_exports_dir() -> str:
    """
    Returns project's default export directory: <root>/data/exports
    __file__ → core/scraper/exporter.py → go up two levels to repo root.
    """
    root = Path(__file__).resolve().parents[2]
    out = root / "data" / "exports"
    out.mkdir(parents=True, exist_ok=True)
    return str(out)

def suggest_filename(fmt: str, scope: str = "All") -> str:
    """
    Generate a timestamped filename like: scraper_all_20250101_120000.csv
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"scraper_{scope.lower()}_{ts}.{fmt.lower()}"
